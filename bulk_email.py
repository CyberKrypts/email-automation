from openpyxl import load_workbook
from termcolor import colored

import base64
import json
import smtplib
import sys

def mail_serv(receiver,message):

    config_file = open('config.py','r')
    config = json.loads(config_file.read())
    config_file.close()
    mail_id     = config['EMAIL_ID']
    password    = config['PASSWORD']

    s = smtplib.SMTP('smtp.gmail.com', 587)

    s.starttls()
    s.login(mail_id, password)
    try:
        s.sendmail(mail_id, receiver, message) 
        print(colored(f'Mail sent to {receiver}','green'))
    except Exception:
        print(colored(f'Mail failed to send {receiver}','red'))
        pass

    s.quit()

def handle_xlsheet(path,msg):

    xl = load_workbook(path)
    ws = xl.active
    max_row = ws.max_row

    for row in range(1, max_row+1):
        email = ws['A'+str(row)].value
        mail_serv(email,msg)

        print(email)
        
if __name__=='__main__':

    if len(sys.argv) == 2:
        path = sys.argv[1]
        config_file = open('config.py','r')
        config = json.loads(config_file.read())
        config_file.close()

        message = """Subject: Cyberkrypts web pen-testing webinar invitation\n\n

                    Hello, This message is from cyberkrypts team. 

Thank you for registering for the free webinar on web pen-testing. This course starts on June 15 and it'll continue for the next 7 days. 
Live classes will be conducted on google meet. The total session duration is 1 hour 30 minutes , 1 hour for the class, and 30 minutes for Q/A. 

course content: https://drive.google.com/file/d/1EzyrmmLTVKDTrpdfrdL0tpX8c0VvjpnP/view?usp=sharing

Join with us in this Whatsapp group: https://chat.whatsapp.com/CGs7Ovuw9aPCAN8gMXuXe2

We will post all updates in this WhatsApp group.



Course instructors : instagram.com/lonewolf.hk
                     instagram.com/comming_soon

Course organizer: instagram.com/comming_soon

                  
                  """

        handle_xlsheet(path,message)
    else:
        print(f"{sys.argv[0]} <path to xlsx file>")